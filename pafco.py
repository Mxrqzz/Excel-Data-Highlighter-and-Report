from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

planilha = load_workbook("pafco.xlsx")

pafco = planilha.active

# Definindo as colunas
assuntoColuna = "C"
cnpjColuna = "D"
dataColuna = "E"
ncmColuna = "I"

# Definindo a data limite da regra
dataRegra = datetime(2024, 2, 19, 23, 59)

# Definindo lista de dados

# REGRAS ATÉ DIA 19/02
assuntoVerdeAntes = []

assuntoAmareloAntes = []

cnpjAntes = []

ncmAntes = []

# REGRAS IGUAL OU DEPOIS DO DIA 20/02
assuntoVerdeApos = []

assuntoAmareloApos = []

cnpjApos = []

ncmApos = []

# Classificando Celulas com cores.

# Alterando Cor da celula da Coluna Assunto
for row in pafco.iter_rows(min_row=3, min_col=3, max_col=5):
    assunto = row[0].value
    data = row[2].value

    if isinstance(data, datetime):
        if data <= dataRegra and str(assunto) in assuntoVerdeAntes:
            row[0].fill = PatternFill(
                start_color="11AD45", end_color="11AD45", fill_type="solid"
            )
        elif data <= dataRegra and str(assunto) in assuntoAmareloAntes:
            row[0].fill = PatternFill(
                start_color="E8E520", end_color="E8E520", fill_type="solid"
            )
        elif data >= dataRegra and str(assunto) in assuntoVerdeApos:
            row[0].fill = PatternFill(
                start_color="11AD45", end_color="11AD45", fill_type="solid"
            )
        elif data >= dataRegra and str(assunto) in assuntoAmareloApos:
            row[0].fill = PatternFill(
                start_color="E8E520", end_color="E8E520", fill_type="solid"
            )

# Alterando Cor da celula da Coluna ncm
for row in pafco.iter_rows(min_row=3, min_col=5, max_col=9):
    data = row[0].value
    ncm = row[4].value

    if isinstance(data, datetime):
        if data <= dataRegra and str(ncm) in ncmAntes:
            row[4].fill = PatternFill(
                start_color="11AD45", end_color="11AD45", fill_type="solid"
            )
        elif data >= dataRegra and str(ncm) in ncmApos:
            row[4].fill = PatternFill(
                start_color="11AD45", end_color="11AD45", fill_type="solid"
            )

# Alterando Cor da celula da Coluna CNPJ
for row in pafco.iter_rows(min_row=3, min_col=4, max_col=5):
    cnpj = row[0].value
    data = row[1].value

    if isinstance(data, datetime):
        if data <= dataRegra and str(cnpj) in cnpjAntes:
            row[0].fill = PatternFill(
                start_color="7C17E8", end_color="7C17E8", fill_type="solid"
            )
        elif data >= dataRegra and str(cnpj) in cnpjApos:
            row[0].fill = PatternFill(
                start_color="7C17E8", end_color="7C17E8", fill_type="solid"
            )


print("Classificação Concluida")

# Definindo Lista de Erros
outros = ["notificação de exigência"]

erros = [
    "não preenchimento do campo",
    "por divergência de informações",
    "código de assunto que não",
    "código de assunto/fato gerador incorreto",
    "código de assunto errado",
]

ausencia = ["ausência de documentação", "insuficiência documental"]

# Realizando o resumo dos Diagnostico
for row in pafco.iter_rows(min_row=3, min_col=13, max_col=14):
    situacao = row[0].value
    descricao = row[1].value

    if situacao == "Indeferida" or situacao == "Em exigência":
        for celulaResumo in outros:
            if celulaResumo in descricao:
                pafco.cell(row=row[0].row, column=15, value="Outro(s)")
                break
        else:
            for celulaResumo in erros:
                if celulaResumo in descricao:
                    pafco.cell(
                        row=row[0].row,
                        column=15,
                        value="Erro de petição/código/informações",
                    )
                    break
            else:
                for celulaResumo in ausencia:
                    if celulaResumo in descricao:
                        pafco.cell(
                            row=row[0].row,
                            column=15,
                            value="Ausência de documento obrigatório",
                        )
                        break

print("Resumo dos Diagnosticos Finalizado")

planilha.save("pafcoFinal.xlsx")

print("Planilha salva")
