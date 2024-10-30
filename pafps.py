from openpyxl import load_workbook
from openpyxl.styles import PatternFill

planilha = load_workbook("arquivoExcel.xlsx")

pafps = planilha.active

# Definindo as colunas
assunto_column = "C"
cnpj_column = "D"
ncm_column = "I"
situacao_column = "M"

# Defininando listas de Dados e codigos
assuntoCanalVerde = []

assuntoCanalAmarelo = []

ncmCanalAmarelo = []

ncmCanalVerde = []

listaExcecoes = []

# Alterando Cor das Celulas

# ALTERANDO COR DAS CELULAS DA COLUNA ASSUNTO:
for celula in pafps[assunto_column]:
    if str(celula.value) in assuntoCanalAmarelo:
        # Definindo cor de fundo da celula
        pafps[celula.coordinate].fill = PatternFill(
            start_color="E8E520", end_color="E8E520", fill_type="solid"
        )
    elif str(celula.value) in assuntoCanalVerde:
        # Definindo a cor de fundo da celula
        pafps[celula.coordinate].fill = PatternFill(
            start_color="11AD45", end_color="11AD45", fill_type="solid"
        )

# ALTERANDO COR DAS CELULAS DA COLUNA CNPJ:
for celula in pafps[cnpj_column]:
    if str(celula.value) in listaExcecoes:
        # Definindo cor de fundo da celula
        pafps[celula.coordinate].fill = PatternFill(
            start_color="7C17E8", end_color="7C17E8", fill_type="solid"
        )

# ALTERANDO COR DAS CELULAS DA COLUNA NCM:
for celula in pafps[ncm_column]:
    if str(celula.value) in ncmCanalAmarelo:
        # Definindo cor de fundo da celula
        pafps[celula.coordinate].fill = PatternFill(
            start_color="E8E520", end_color="E8E520", fill_type="solid"
        )
    elif str(celula.value) in ncmCanalVerde:
        # Definindo cor de fundo da celula
        pafps[celula.coordinate].fill = PatternFill(
            start_color="11AD45", end_color="11AD45", fill_type="solid"
        )

print("Classificação dos Assuntos, CNPjs e NCMs Finalizada.")

# Definindo Lista de Erros
outros = ["notificação de exigência"]

erros = [
    "não preenchimento do campo",
    "por divergência de informações",
    "código de assunto que não",
    "código de assunto/fato gerador incorreto",
    "código de assunto errado",
]

ausencia = ["ausência de documentação", "insuficiência documental"]

# Realizando o resumo dos Diagnostico
for row in pafps.iter_rows(min_row=3, min_col=13, max_col=14):
    situacao = row[0].value
    descricao = row[1].value

    if situacao == "Indeferida":
        for celulaResumo in outros:
            if celulaResumo in descricao:
                pafps.cell(row=row[0].row, column=15, value="Outro(s)")
                break
        else:  # se não encontrar na lista outros procura na lista erros
            for celulaResumo in erros:
                if celulaResumo in descricao:
                    pafps.cell(
                        row=row[0].row,
                        column=15,
                        value="Erro de petição/código/informações",
                    )
                    break
            else:  # se não encontrar na erros outros procura na lista ausencia
                for celulaResumo in ausencia:
                    if celulaResumo in descricao:
                        pafps.cell(
                            row=row[0].row,
                            column=15,
                            value="Ausência de documento obrigatório",
                        )
                        break

print("Resumo dos Diagnosticos Finalizado")

planilha.save("arquivoExcelFinal.xlsx")

print("Planilha Finalizada.")
