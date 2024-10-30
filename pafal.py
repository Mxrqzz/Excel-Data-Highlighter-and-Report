from openpyxl import load_workbook
from openpyxl.styles import PatternFill

planilha = load_workbook("pafal.xlsx")

pafal = planilha.active

assunto_column = "C"
cnpj_column = "D"
ncm_column = "I"

# Criando listas com codigos para assunto, ncm e cnpj
assuntoAmarelo = []

ncmVerde = []

ncmAmarelo = []

cnpj = []

# Realizando a Classificação por cores nas celulas

for celula in pafal[assunto_column]:
    if str(celula.value) in assuntoAmarelo:
        celula.fill = PatternFill(
            start_color="E8E520", end_color="E8E520", fill_type="solid"
        )

print("Classificação da Coluna Assunto finalizada")

for celula in pafal[ncm_column]:
    if str(celula.value) in ncmAmarelo:
        celula.fill = PatternFill(
            start_color="E8E520", end_color="E8E520", fill_type="solid"
        )
    elif str(celula.value) in ncmVerde:
        celula.fill = PatternFill(
            start_color="11AD25", end_color="11AD25", fill_type="solid"
        )

print("Classificação da Coluna NCM finalizada")

for celula in pafal[cnpj_column]:
    if str(celula.value) in cnpj:
        celula.fill = PatternFill(
            start_color="7C17E8", end_color="7C17E8", fill_type="solid"
        )

print("Classificação da Coluna CNPJ finalizada")

# Definindo Lista de Erros
outros = ["notificação de exigência"]

erros = [
    "não preenchimento do campo",
    "por divergência de informações",
    "código de assunto que não",
    "código de assunto/fato gerador incorreto",
    "código de assunto errado",
]

ausencia = ["ausência de documentação", "insuficiência documental"]

# Realizando o resumo dos Diagnostico
for row in pafal.iter_rows(min_row=3, min_col=13, max_col=14):
    situacao = row[0].value
    descricao = row[1].value

    if situacao == "Indeferida" or situacao == "Em exigência":
        for resumo in outros:
            if resumo in descricao:
                pafal.cell(row=row[0].row, column=15, value="Outro(s)")
                break
        else:
            for resumo in erros:
                if resumo in descricao:
                    pafal.cell(
                        row=row[0].row,
                        column=15,
                        value="Erro de Petição/código/informações",
                    )
                    break
            else:
                for resumo in ausencia:
                    if resumo in descricao:
                        pafal.cell(
                            row=row[0].row,
                            column=15,
                            value="Ausência de documento obrigatório",
                        )
                        break

print("Resumo Realizado")

planilha.save("PafalFinal.xlsx")

print("finalizado")
