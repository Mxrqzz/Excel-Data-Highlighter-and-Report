from openpyxl import load_workbook
from openpyxl.styles import PatternFill

planilha = load_workbook("ArquivoExcel.xlsx")

pafme = planilha.active

# Definindo as colunas
assunto_column = "C"
cnpj_column = "D"
situacao_column = "M"

# Defininando listas de Dados e codigos
assuntoVermelho = []

assuntoAmarelo = []

assuntoVerde = []

listaCnpjs = []

for celula in pafme[assunto_column]:
    if str(celula.value) in assuntoVermelho:
        pafme[celula.coordinate].fill = PatternFill(
            start_color="00FF0000", end_color="00FF0000", fill_type="solid"
        )
    elif str(celula.value) in assuntoVerde:
        pafme[celula.coordinate].fill = PatternFill(
            start_color="11AD25", end_color="11AD25", fill_type="solid"
        )
    elif str(celula.value) in assuntoAmarelo:
        pafme[celula.coordinate].fill = PatternFill(
            start_color="E8E520", end_color="E8E520", fill_type="solid"
        )

print("classificação Assunto concluida")

for celula in pafme[cnpj_column]:
    if str(celula.value) in listaCnpjs:
        pafme[celula.coordinate].fill = PatternFill(
            start_color="7C17E8", end_color="7C17E8", fill_type="solid"
        )

print("classificação CNPJ concluida")

outros = ["notificação de exigência"]

erros = [
    "não preenchimento do campo",
    "por divergência de informações",
    "código de assunto que não",
    "código de assunto/fato gerador incorreto",
    "código de assunto errado",
]

ausencia = ["ausência de documentação", "insuficiência documental"]

# Realizando o resumo dos Diagnostico
for row in pafme.iter_rows(min_row=3, min_col=13, max_col=14):
    situacao = row[0].value
    descricao = row[1].value

    if situacao == "Indeferida":
        for celulaResumo in outros:
            if celulaResumo in descricao:
                pafme.cell(row=row[0].row, column=15, value="Outro(s)")
                break
        else:
            for celulaResumo in erros:
                if celulaResumo in descricao:

                    pafme.cell(
                        row=row[0].row,
                        column=15,
                        value="Erro de petição/código/informações",
                    )
                    break
            else:
                for celulaResumo in ausencia:
                    if celulaResumo in descricao:

                        pafme.cell(
                            row=row[0].row,
                            column=15,
                            value="Ausência de documento obrigatório",
                        )
                        break

print("Resumo dos Diagnosticos Finalizado")

planilha.save("arquivoExcelFinal.xlsx")

print("planilha finalizada")
